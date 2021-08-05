from NaverNewsCrawler import NaverNewsCrawler

####사용자로 부터 기사 수집을 원하는 키워드를 input을 이용해 입력받아 ? 부분에 넣으세요

keyword = input("기사 수집을 원하는 키워드를 입력해주세요: ")
crawler = NaverNewsCrawler(keyword)

#### 수집한 데이터를 저장할 엑셀 파일명을 input을 이용해 입력받아 ? 부분에 넣으세요

news_excel = input("수집한 데이터를 저장할 엑셀 파일명을 입력해주세요: ")
crawler.get_news(news_excel)

#### 아래코드를 실행해 이메일 발송 기능에 필요한 모듈을 임포트하세요.
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import smtplib
import re

#### gmail 발송 기능에 필요한 계정 정보를 아래 코드에 입력하세요.
SMTP_SERVER = 'smtp.gmail.com'
SMTP_PORT = 465
SMTP_USER = '' #테스트 완료 후 삭제하여 깃허브 반영
SMTP_PASSWORD = '' #테스트 완료 후 삭제하여 깃허브 반영

#### 아래 코드를 실행해 메일 발송에 필요한 send_mail 함수를 만드세요.
def send_mail(name, addr, subject, contents, attachment=None):
    if not re.match('(^[a-zA-Z0-9_.-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$)', addr):
        print('Wrong email')
        return

    msg = MIMEMultipart('alternative')
    if attachment:
        msg = MIMEMultipart('mixed')

    msg['From'] = SMTP_USER
    msg['To'] = addr
    msg['Subject'] = name + '님, ' + subject

    text = MIMEText(contents, _charset='utf-8')
    msg.attach(text)

    if attachment:

        from email.mime.base import MIMEBase
        from email import encoders

        file_data = MIMEBase('application', 'octect-stream')
        file_data.set_payload(open(attachment, 'rb').read())
        encoders.encode_base64(file_data)

        import os
        filename = os.path.basename(attachment)
        file_data.add_header('Content-Disposition', 'attachment; filename="' + filename + '"')
        msg.attach(file_data)

    smtp = smtplib.SMTP_SSL(SMTP_SERVER, SMTP_PORT)
    smtp.login(SMTP_USER, SMTP_PASSWORD)
    smtp.sendmail(SMTP_USER, addr, msg.as_string())
    smtp.quit()

#### 프로젝트 폴더에 있는 email_list.xlsx 파일에 이메일 받을 사람들의 정보를 입력하세요.

from openpyxl import Workbook

wb = Workbook()
ws = wb.active

ws['A1'] = '번호'
ws['B1'] = '이름'
ws['C1'] = '이메일'

ws['A2'] = '1'
ws['B2'] = '' #테스트 완료 후 삭제하여 깃허브 반영
ws['C2'] = '' #테스트 완료 후 삭제하여 깃허브 반영

wb.save('email_list.xlsx')

#### 엑셀 파일의 정보를 읽어올 수 있는 모듈을 import하세요.

from openpyxl import load_workbook

#### email_list.xlsx 파일을 읽어와 해당 사람들에게 수집한 뉴스 정보 엑셀 파일을 send_mail 함수를 이용해 전송하세요.

e_xlsx = load_workbook('email_list.xlsx', data_only=True)
data = e_xlsx.active

emails = []

for row in data.iter_rows(max_row=2):
    peoplelist = []
    for cell in row:
        peoplelist.append(cell.value)
    emails.append(peoplelist)

name = (emails[1][1])
address = (emails[1][2])
subject = keyword + '뉴스 입니다'
contents = keyword + '뉴스 요약본은 파일로 첨부하여 보냅니다'
file = keyword + '.xlsx'

send_mail(name, address, subject, contents, file)
